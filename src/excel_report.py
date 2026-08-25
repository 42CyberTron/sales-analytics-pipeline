import os
import sqlite3
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(BASE_DIR, "db", "sales.db")
OUTPUTS_DIR = os.path.join(BASE_DIR, "outputs")
METRICS_PATH = os.path.join(OUTPUTS_DIR, "model_metrics.json")
FORECAST_CSV = os.path.join(BASE_DIR, "data", "processed", "forecast_output.csv")

# Styles
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
KPI_FONT = Font(name="Calibri", bold=True, size=16, color="2F5496")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F3864")
CURRENCY_FMT = '#,##0.00'
PCT_FMT = '0.00%'
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def get_db():
    return sqlite3.connect(DB_PATH)


def style_header_row(ws, row_num=1, max_col=10):
    """Apply header styling."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def auto_width(ws, min_width=10, max_width=30):
    """Auto-size columns based on content."""
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def build_summary_tab(wb):
    """Tab 1: KPIs with conditional formatting."""
    ws = wb.create_sheet("Summary")
    conn = get_db()

    # Query KPIs
    total_revenue = pd.read_sql(
        "SELECT ROUND(SUM(revenue_gbp), 2) AS val FROM transactions", conn
    ).iloc[0]["val"]

    top_country = pd.read_sql("""
        SELECT c.country, ROUND(SUM(t.revenue_gbp), 2) AS rev
        FROM transactions t JOIN customers c ON t.customer_id = c.customer_id
        GROUP BY c.country ORDER BY rev DESC LIMIT 1
    """, conn)

    mom_data = pd.read_sql("""
        WITH monthly AS (
            SELECT substr(invoice_date, 1, 7) AS month, SUM(revenue_gbp) AS revenue
            FROM transactions GROUP BY month
        )
        SELECT month, revenue,
               LAG(revenue) OVER (ORDER BY month) AS prev
        FROM monthly ORDER BY month
    """, conn)

    latest_mom = (
        (mom_data["revenue"].iloc[-1] - mom_data["prev"].iloc[-1])
        / mom_data["prev"].iloc[-1]
        if mom_data["prev"].iloc[-1] and mom_data["prev"].iloc[-1] > 0 else 0
    )
    conn.close()

    # Load model metrics
    if os.path.exists(METRICS_PATH):
        with open(METRICS_PATH) as f:
            metrics = json.load(f)
        forecast_next = metrics.get("prophet_forecast_next_month", 0)
        improvement = metrics.get("improvement_pct", 0)
    else:
        forecast_next = 0
        improvement = 0

    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = "Sales Analytics Dashboard — Online Retail II"
    ws["A1"].font = TITLE_FONT

    ws["A2"] = f"Generated: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, size=9, color="808080")

    # KPI section
    kpi_start = 4
    ws.cell(row=kpi_start, column=1, value="Key Performance Indicators").font = Font(
        bold=True, size=12, color="1F3864"
    )

    kpis = [
        ("Total Revenue (GBP)", total_revenue, "currency"),
        ("Top Country", f"{top_country.iloc[0]['country']} (£{top_country.iloc[0]['rev']:,.2f})", "text"),
        ("MoM Growth (Latest)", latest_mom, "pct"),
        ("Forecast Next Month (GBP)", forecast_next, "currency"),
        ("Model Improvement vs Baseline", f"{improvement}%", "text"),
    ]

    for i, (label, value, fmt) in enumerate(kpis):
        row = kpi_start + 1 + i
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=11)
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = KPI_FONT
        cell.border = THIN_BORDER

        if fmt == "currency":
            cell.number_format = CURRENCY_FMT
        elif fmt == "pct":
            cell.number_format = PCT_FMT
            # Conditional formatting
            if value >= 0:
                cell.fill = GREEN_FILL
            else:
                cell.fill = RED_FILL

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 35

    return ws


def build_monthly_trend_tab(wb):
    """Tab 2: Monthly trend table + chart."""
    ws = wb.create_sheet("Monthly Trend")
    conn = get_db()

    monthly = pd.read_sql("""
        SELECT substr(invoice_date, 1, 7) AS month,
               ROUND(SUM(revenue_gbp), 2) AS revenue,
               COUNT(DISTINCT invoice_no) AS orders
        FROM transactions
        GROUP BY month ORDER BY month
    """, conn)
    conn.close()

    ws["A1"] = "Monthly Revenue Trend"
    ws["A1"].font = TITLE_FONT

    # Write headers
    headers = ["Month", "Revenue (GBP)", "Orders"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, row_num=3, max_col=3)

    # Write data
    for i, row in monthly.iterrows():
        r = 4 + i
        ws.cell(row=r, column=1, value=row["month"])
        ws.cell(row=r, column=2, value=row["revenue"]).number_format = CURRENCY_FMT
        ws.cell(row=r, column=3, value=row["orders"])

    data_end = 3 + len(monthly)

    # Chart
    chart = LineChart()
    chart.title = "Monthly Revenue Trend"
    chart.y_axis.title = "Revenue (GBP)"
    chart.x_axis.title = "Month"
    chart.style = 10
    chart.height = 12
    chart.width = 20

    data_ref = Reference(ws, min_col=2, min_row=3, max_row=data_end)
    cats = Reference(ws, min_col=1, min_row=4, max_row=data_end)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "E3")

    auto_width(ws)
    ws.freeze_panes = "A4"

    return ws


def build_forecast_tab(wb):
    """Tab 3: Forecast vs actual with metrics."""
    ws = wb.create_sheet("Forecast")

    ws["A1"] = "Revenue Forecast — Holt-Winters vs Baseline"
    ws["A1"].font = TITLE_FONT

    # Metrics display
    if os.path.exists(METRICS_PATH):
        with open(METRICS_PATH) as f:
            metrics = json.load(f)

        ws["A3"] = "Model Performance"
        ws["A3"].font = Font(bold=True, size=12, color="1F3864")

        metric_rows = [
            ("Model", metrics.get("model", "N/A")),
            ("Baseline RMSE", metrics.get("baseline_rmse", 0)),
            ("Holt-Winters RMSE", metrics.get("hw_rmse", 0)),
            ("Improvement", f"{metrics.get('improvement_pct', 0)}%"),
        ]
        for i, (label, val) in enumerate(metric_rows):
            ws.cell(row=4 + i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=4 + i, column=2, value=val)

    # Forecast table
    if os.path.exists(FORECAST_CSV):
        forecast = pd.read_csv(FORECAST_CSV)
        start_row = 10

        ws.cell(row=start_row, column=1, value="Forecast Detail").font = Font(
            bold=True, size=12, color="1F3864"
        )

        headers = ["Month", "Forecast (GBP)", "Is Future"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=start_row + 1, column=col, value=h)
        style_header_row(ws, row_num=start_row + 1, max_col=3)

        for i, row in forecast.iterrows():
            r = start_row + 2 + i
            ws.cell(row=r, column=1, value=str(row["month"])[:10])
            ws.cell(row=r, column=2, value=round(row["forecast_gbp"], 2)).number_format = CURRENCY_FMT
            ws.cell(row=r, column=3, value="Yes" if row["is_future"] else "No")

        # Chart
        data_end = start_row + 1 + len(forecast)
        chart = LineChart()
        chart.title = "Holt-Winters Forecast"
        chart.y_axis.title = "Revenue (GBP)"
        chart.style = 10
        chart.height = 12
        chart.width = 22

        data_ref = Reference(ws, min_col=2, min_row=start_row + 1, max_row=data_end)
        cats = Reference(ws, min_col=1, min_row=start_row + 2, max_row=data_end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "F3")

    auto_width(ws)
    ws.freeze_panes = "A4"

    return ws


def build_raw_data_tab(wb):
    """Tab 4: Raw data for auditability."""
    ws = wb.create_sheet("Top Products")
    conn = get_db()

    top_products = pd.read_sql("""
        SELECT p.stock_code, p.description,
               ROUND(SUM(t.revenue_gbp), 2) AS total_revenue,
               SUM(t.quantity) AS total_units,
               COUNT(DISTINCT t.invoice_no) AS num_orders
        FROM transactions t
        JOIN products p ON t.stock_code = p.stock_code
        GROUP BY p.stock_code, p.description
        ORDER BY total_revenue DESC
        LIMIT 50
    """, conn)
    conn.close()

    ws["A1"] = "Top 50 Products by Revenue"
    ws["A1"].font = TITLE_FONT

    headers = list(top_products.columns)
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, row_num=3, max_col=len(headers))

    for i, row in top_products.iterrows():
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=4 + i, column=col, value=val)
            if col == 3:
                cell.number_format = CURRENCY_FMT

    # Chart
    data_end = 3 + len(top_products)
    chart = BarChart()
    chart.type = "col"
    chart.title = "Top 10 Products by Revenue"
    chart.y_axis.title = "Revenue (GBP)"
    chart.style = 10
    chart.height = 12
    chart.width = 20

    data_ref = Reference(ws, min_col=3, min_row=3, max_row=min(13, data_end))
    cats = Reference(ws, min_col=2, min_row=4, max_row=min(13, data_end))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "F3")

    auto_width(ws)
    ws.freeze_panes = "A4"

    return ws


if __name__ == "__main__":
    os.makedirs(OUTPUTS_DIR, exist_ok=True)

    print("[EXCEL] Building multi-tab report...")
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    build_summary_tab(wb)
    print("  Tab 1: Summary (KPIs + conditional formatting)")

    build_monthly_trend_tab(wb)
    print("  Tab 2: Monthly Trend (table + line chart)")

    build_forecast_tab(wb)
    print("  Tab 3: Forecast (metrics + CI chart)")

    build_raw_data_tab(wb)
    print("  Tab 4: Top Products (table + bar chart)")

    out_path = os.path.join(OUTPUTS_DIR, "report.xlsx")
    wb.save(out_path)
    print(f"\n[EXCEL] Report saved to: {out_path}")
